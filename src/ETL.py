"""
Descripción del módulo

Esté módulo realiza el proceso para poder generar 
las proyecciones de volumen y ventas con base al historico
de ventas reales. 

Autor: [Cristian Segura]
Fecha: [16/11/2023]
"""
# ----------------------------------------------
# Segmento 1: Importar bibliotecas
# ----------------------------------------------
import os  #Libreria para listar archivos de una ruta
import time 
import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font 
# ----------------------------------------------


# ----------------------------------------------
# Inicio del temporizador
tiempo_inicio = time.time()
print('---------------------------------\n')
print(' Bienvenido a la app Para realizar la proyección de Ventas\n')
print('---------------------------------')
print(' **  Se esta tomando la data de Ventas Historicas Reales ** ')
# ----------------------------------------------


# ----------------------------------------------
# Segmento 2: Lectura de archivos y asignación a dataframes
# ----------------------------------------------
#----  Base Ventas Reales
# Especifica las columnas que deseas que sean de tipo texto
Base_Texto = ['COD ARTICULO','Mes']
BaseVentasReal = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/VentasReales/Base Real Historica.xlsx'
                               , sheet_name='Base'
                               , header=0
                               , names=None
                               , index_col = None
                               , usecols= 'A:AB'
                               , engine= 'openpyxl'
                               , dtype={col: str for col in Base_Texto}) #---- Cast to type text
BaseVentasReal['COD ARTICULO SIN EXTRAC'] = None #---- Creación de columna vacia
BaseVentasReal['ARTICULO SIN EXTRAC'] = None #---- Creación de columna vacia
df_BaseVentasReal = pd.DataFrame(BaseVentasReal)

#----   Base Sin extra contenido para actualizar Ventas reales
BdSinExtra_Texto = ['CODI_ARTICULO EXTRAC','COD ARTICULO REG']
BdSinExtra = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/VentasReales/Base Real Historica.xlsx'
                           , sheet_name='Art Extrac'
                           , header=0
                           , names=None
                           , index_col = None
                           , usecols= 'E:H'
                           , engine= 'openpyxl'
                           , dtype={col: str for col in BdSinExtra_Texto})#---- Cast to type text
df_BdSinExtra = pd.DataFrame(BdSinExtra)
#print(df_BaseVentasReal)
#print(df_BdSinExtra)
# ----------------------------------------------


# ----------------------------------------------
# Segmento 3: Creación de funciones para transformaciones
# ----------------------------------------------
# Función personalizada para Asignar el código articulo sin extracontenido cuando aplique
# ---
def fn_CodSinExtra(row) :
    if row['TIPO_OFERTA'] == 'OS2':
        # Lógica para la primer condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] == 
                                     row['COD ARTICULO']]['COD ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    elif row['TIPO_OFERTA'] == 'K05':
        # Lógica para la segunda condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] ==
                                     row['COD ARTICULO']]['COD ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    # Retornar el valor original si no se cumple ninguna de las condiciones
    return row['COD ARTICULO']
# ---
# Crear una función personalizada para Asignar
# la descripción del articulo sin extracontenido cuando aplique
# ---
def fn_DescSinExtra(row):
    if row['TIPO_OFERTA'] == 'OS2':
        # Lógica para la primer condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] ==
                                     row['COD ARTICULO']]['DESC_ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    elif row['TIPO_OFERTA'] == 'K05':
        # Lógica para la segunda condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] ==
                                     row['COD ARTICULO']]['DESC_ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    # Retornar el valor original si no se cumple ninguna condición
    return row['ARTICULO']
# ----------------------------------------------


# ----------------------------------------------
# Segmento 4: Columnas Añadidas con funciones y llaves
# ----------------------------------------------
# Aplicar la función a cada fila del DataFrame
df_BaseVentasReal['COD ARTICULO SIN EXTRAC'] = df_BaseVentasReal.apply(fn_CodSinExtra, axis=1)
df_BaseVentasReal['ARTICULO SIN EXTRAC'] = df_BaseVentasReal.apply(fn_DescSinExtra, axis=1)
# ----------------------------------------------


# ----------------------------------------------
# Segmento 5: Control de historico con N cantidad de meses
# para calcular la proyección del Volumen
# ----------------------------------------------
#----  Copiar data del dataframe de VentasReal
df_HisVolumen = df_BaseVentasReal.copy()
# Convertir las columnas 'año' y 'mes' a tipo str y luego a tipo int
df_HisVolumen['Año'] = df_HisVolumen['Año'].astype(str).astype(int)
df_HisVolumen['Mes'] = df_HisVolumen['Mes'].astype(str).astype(int)
# Calcular la fecha ficticia como el primer día de cada mes
df_HisVolumen['fecha'] = pd.to_datetime(df_HisVolumen['Año'].astype(str) + '-' + df_HisVolumen['Mes'].astype(str) + '-01')
#----
# Variable para almacenar el número de mes
MES_HISTVOLUMEN = None
# Bucle para capturar un mes valido
while MES_HISTVOLUMEN is None:
    try:
        # Solicitar al usuario que ingrese el mes
        numero = int(input(" Ingrese la cantidad de meses del historico para Proyectar el Volumen - "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= numero <= 12:
            MES_HISTVOLUMEN = numero  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese un número en el rango del 1 al 12.")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")
# Imprimir el número válido
print("\n Se Calculará para los Ultimos: ", MES_HISTVOLUMEN, " Meses")
#----
# Calcular la fecha de los ultimos N meses desde la fecha máxima
fecha_maximaVolumen = df_HisVolumen['fecha'].max()
fecha_limiteVolumen = fecha_maximaVolumen - relativedelta(months=(MES_HISTVOLUMEN)-1)
# Filtrar el DataFrame por los últimos 3 meses
df_filtradoVolumen = df_HisVolumen[df_HisVolumen['fecha'] >= fecha_limiteVolumen].copy()
# Eliminar la columna 'fecha' si no es necesaria
df_filtradoVolumen = df_filtradoVolumen.sort_values('fecha', ascending=False)
df_filtradoVolumen = df_filtradoVolumen.drop(columns=['fecha'])
#libroMes.close()
# Imprimir el DataFrame resultante para validar
# que se tomo el historico correcto
df_filtradoVolumen.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/4) Validación/ValidaciónMesVolumen.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
# ----------------------------------------------


# ----------------------------------------------
# Segmento 6: Porcentaje Individual y homologación tabla dinamica
# ----------------------------------------------
#---- Columnas que se requieren unicamente
columnas_tb_dinamica1 = ['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                         , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                         , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                         , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                         , 'NOMBRE_TIPOOFERTA', 'KILO_NETO']
df_tabla_dinamica1 = df_filtradoVolumen[columnas_tb_dinamica1]
#---- Agrupa por las columnas seleccionadas y suma la columna 'KILO_NETO'
df_agrupado1 = df_tabla_dinamica1.groupby(['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                                           , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                                           , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                                           , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                                           , 'NOMBRE_TIPOOFERTA']).sum().reset_index()
#---- Calcular el total de 'KILO_NETO' agrupado por 'NOMBRE_CATEGORIA'
total_por_categoria = df_agrupado1.groupby('NOMBRE_CATEGORIA')['KILO_NETO'].sum().reset_index()
total_por_categoria.columns = ['NOMBRE_CATEGORIA'
                            , 'KILONETOCATEGORIA']

#---- Fusionar df_agrupado1 con el total por categoría
df_agrupado1 = pd.merge(df_agrupado1
                        , total_por_categoria
                        , on='NOMBRE_CATEGORIA')
#---- Calcular la nueva columna como el porcentaje del total 'KILO_NETO'
df_agrupado1['PORC_TOTALCATEGORIA'] = (df_agrupado1['KILO_NETO']
                                       / df_agrupado1['KILONETOCATEGORIA'])

df_agrupado1.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenProyectado.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
# ----------------------------------------------

#"""
# ----------------------------------------------
# Segmento 7: Diseño Visual del Excel con openpyxl
# ----------------------------------------------
# Ruta al archivo
archivo_xlsx = 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenProyectado.xlsx'
# Cargar el libro de trabajo
wb = load_workbook(archivo_xlsx)
hoja = wb['Hoja1']

# Aplica formato al encabezado (primera fila)
for cell in hoja[1]:
    cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)

# Ajustar el ancho de las columnas al contenido
for columna in hoja.columns:
    max_length = 0
    columna = [cell for cell in columna]
    for cell in columna:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 4)
    hoja.column_dimensions[columna[0].column_letter].width = adjusted_width
# Guarda el archivo Excel
wb.save(archivo_xlsx)
# Cierra el libro de trabajo
wb.close()
# ----------------------------------------------
#"""

# ----------------------------------------------
# Segmento 8: Tiempos de ejecución
# ----------------------------------------------
# Calcular la duración total de la ejecución
tiempo_fin = time.time()
# Calcular la duración total de la ejecución en segundos
duracion_total_segundos = tiempo_fin - tiempo_inicio
# Obtener horas, minutos y segundos
horas, rem = divmod(duracion_total_segundos, 3600)
minutos, segundos = divmod(rem, 60)
# Imprimir el tiempo total de ejecución en formato hh/mm/ss
print(f' Tiempo total de ejecución: {int(horas):02d}:{int(minutos):02d}:{int(segundos):02d}')
#time.sleep(10)
input(" Presiona Enter para salir")
# ----------------------------------------------