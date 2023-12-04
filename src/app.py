"""
Descripción del módulo

Esté módulo realiza el proceso para poder generar 
las proyecciones de volumen y ventas con base al historico
de ventas reales. 

Autor: [Cristian Segura]
Fecha: [16/11/2023]
"""
# ----------------------------------------------
# Segmento 1: Importar bibliotecas
# ----------------------------------------------
import os  #Libreria para listar archivos de una ruta
import time
import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
# ----------------------------------------------


# ----------------------------------------------
# Inicio del temporizador
tiempo_inicio = time.time()
print('---------------------------------\n')
print(' Bienvenido a la app Para realizar la proyección de Ventas\n')
print('---------------------------------')
print(' **  Se esta tomando la data de Ventas Historicas Reales ** ')
# ----------------------------------------------


# ----------------------------------------------
# Segmento 2: Lectura de archivos y asignación a dataframes
# ----------------------------------------------
#----  Base Ventas Reales
# Especifica las columnas que deseas que sean de tipo
Base_Texto = ['COD ARTICULO','Mes']
BaseVentasReal = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/VentasReales/Base Real Historica.xlsx'
                               , sheet_name='Base'
                               , header=0
                               , names=None
                               , index_col = None
                               , usecols= 'A:AB'
                               , engine= 'openpyxl'
                               , dtype={col: str for col in Base_Texto}) #---- Cast to type text
BaseVentasReal['COD ARTICULO SIN EXTRAC'] = None #---- Creación de columna vacia
BaseVentasReal['ARTICULO SIN EXTRAC'] = None #---- Creación de columna vacia
df_BaseVentasReal = pd.DataFrame(BaseVentasReal)

#----   Base Sin extra contenido para actualizar Ventas reales
BdSinExtra_Texto = ['CODI_ARTICULO EXTRAC','COD ARTICULO REG']
BdSinExtra = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/VentasReales/Base Real Historica.xlsx'
                           , sheet_name='Art Extrac'
                           , header=0
                           , names=None
                           , index_col = None
                           , usecols= 'E:H'
                           , engine= 'openpyxl'
                           , dtype={col: str for col in BdSinExtra_Texto})#---- Cast to type text
df_BdSinExtra = pd.DataFrame(BdSinExtra)
#print(df_BaseVentasReal)
#print(df_BdSinExtra)
# ----------------------------------------------


# ----------------------------------------------
# Segmento 3: Creación de funciones para transformaciones
# ----------------------------------------------
# Función personalizada para Asignar el código 
# articulo sin extracontenido cuando aplique
# ---
def fn_CodSinExtra(row) :
    if row['TIPO_OFERTA'] == 'OS2':
        # Lógica para la primer condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] ==
                                     row['COD ARTICULO']]['COD ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    elif row['TIPO_OFERTA'] == 'K05':
        # Lógica para la segunda condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] ==
                                     row['COD ARTICULO']]['COD ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    # Retornar el valor original si no se cumple ninguna condición
    return row['COD ARTICULO']
# ---
# Crear una función personalizada para Asignar
# la descripción del articulo sin extracontenido cuando aplique
# ---
def fn_DescSinExtra(row):
    if row['TIPO_OFERTA'] == 'OS2':
        # Lógica para la primer condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] ==
                                     row['COD ARTICULO']]['DESC_ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    elif row['TIPO_OFERTA'] == 'K05':
        # Lógica para la segunda condición y el merge correspondiente
        merge_result = df_BdSinExtra[df_BdSinExtra['CODI_ARTICULO EXTRAC'] ==
                                     row['COD ARTICULO']]['DESC_ARTICULO REG'].values
        if len(merge_result) > 0:
            return merge_result[0]
    # Retornar el valor original si no se cumple ninguna condición
    return row['ARTICULO']
# ----------------------------------------------


# ----------------------------------------------
# Segmento 4: Columnas Añadidas con funciones y llaves
# ----------------------------------------------
# Aplicar la función a cada fila del DataFrame
df_BaseVentasReal['COD ARTICULO SIN EXTRAC'] = df_BaseVentasReal.apply(fn_CodSinExtra, axis=1)
df_BaseVentasReal['ARTICULO SIN EXTRAC'] = df_BaseVentasReal.apply(fn_DescSinExtra, axis=1)
# ---
# ----------------------------------------------


# ----------------------------------------------
# Segmento 5: Control de historico con N cantidad de meses
# para calcular la proyección del Volumen
# ----------------------------------------------
#----  Copiar data del dataframe de VentasReal
df_HisVolumen = df_BaseVentasReal.copy()
# Convertir las columnas 'año' y 'mes' a tipo str y luego a tipo int
df_HisVolumen['Año'] = df_HisVolumen['Año'].astype(str).astype(int)
df_HisVolumen['Mes'] = df_HisVolumen['Mes'].astype(str).astype(int)
# Calcular la fecha ficticia como el primer día de cada mes
df_HisVolumen['fecha'] = pd.to_datetime(df_HisVolumen['Año'].astype(str) + '-' + df_HisVolumen['Mes'].astype(str) + '-01')
#----
# Variable para almacenar el número de mes
MES_HISTVOLUMEN = None
# Bucle para capturar un mes valido
while MES_HISTVOLUMEN is None:
    try:
        # Solicitar al usuario que ingrese el mes
        numero = int(input(" Ingrese la cantidad de meses del historico para Proyectar el Volumen - "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= numero <= 12:
            MES_HISTVOLUMEN = numero  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese un número en el rango del 1 al 12.")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")
# Imprimir el número válido
print("\n Se Calculará para los Ultimos: ", MES_HISTVOLUMEN, " Meses")
#----
# Calcular la fecha de los ultimos N meses desde la fecha máxima
fecha_maximaVolumen = df_HisVolumen['fecha'].max()
fecha_limiteVolumen = fecha_maximaVolumen - relativedelta(months=(MES_HISTVOLUMEN)-1)
# Filtrar el DataFrame por los últimos 3 meses
df_filtradoVolumen = df_HisVolumen[df_HisVolumen['fecha'] >= fecha_limiteVolumen].copy()
# Eliminar la columna 'fecha' si no es necesaria
df_filtradoVolumen = df_filtradoVolumen.sort_values('fecha', ascending=False)
df_filtradoVolumen = df_filtradoVolumen.drop(columns=['fecha'])
#libroMes.close()
# Imprimir el DataFrame resultante para validar
# que se tomo el historico correcto
df_filtradoVolumen.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/Validación/ValidaciónMesVolumen.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
# ----------------------------------------------


# ----------------------------------------------
# Segmento 6: Porcentaje Individual y homologación tabla dinamica
# ----------------------------------------------
#---- Tablas que se requieren unicamente
columnas_tb_dinamica1 = ['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                         , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                         , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                         , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                         , 'NOMBRE_TIPOOFERTA', 'KILO_NETO']
df_tabla_dinamica1 = df_filtradoVolumen[columnas_tb_dinamica1]
#---- Agrupa por las columnas seleccionadas y suma la columna 'KILO_NETO'
df_agrupado1 = df_tabla_dinamica1.groupby(['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                                           , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                                           , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                                           , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                                           , 'NOMBRE_TIPOOFERTA']).sum().reset_index()
#---- Calcular el total de 'KILO_NETO' agrupado por 'NOMBRE_CATEGORIA'
total_por_categoria = df_agrupado1.groupby('NOMBRE_CATEGORIA')['KILO_NETO'].sum().reset_index()
total_por_categoria.columns = ['NOMBRE_CATEGORIA'
                            , 'KILONETOCATEGORIA']
df_agrupado1.to_excel('C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenProyectado111.xlsx'
                    , index=False
                    , sheet_name='Hoja1'
                    , engine='openpyxl')
#---- Fusionar df_agrupado1 con el total por categoría
df_agrupado1 = pd.merge(df_agrupado1
                        , total_por_categoria
                        , on='NOMBRE_CATEGORIA')
#---- Calcular la nueva columna como el porcentaje del total 'KILO_NETO'
df_agrupado1['PORC_TOTALCATEGORIA'] = (df_agrupado1['KILO_NETO']
                                       / df_agrupado1['KILONETOCATEGORIA'])
# ----------------------------------------------


# ----------------------------------------------
# Segmento 7: Proceso de explosión de Volumenes
# ----------------------------------------------
# Validación del archivo a valorizar para cargar en un dataframe
# --- Ruta del directorio
RUTA = 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenDeProyección'
# Nombre de los archivos que esperamos encontrar
BDKILOSVAL = 'bdKilosVal.xlsx'
BDCATEGORIA = 'bdCategoria.xlsx'
BDCANAL = 'bdCanal.xlsx'

while True:
    # Lista los archivos en la ruta
    FILES = os.listdir(RUTA)

    # Verifica qué archivo está presente
    if BDKILOSVAL in FILES:
        print('\n---------------------------------\n')
        print(f" El archivo {BDKILOSVAL} está presente. Realizar función para Ventas.")
        # Aquí puedes agregar la lógica específica para el archivo Ventas bdKilosVal.xlsx
        bdKilosVal = pd.read_excel(
            io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenDeProyección/bdKilosVal.xlsx'
                                    , sheet_name='Hoja1'
                                    , header=0
                                    , names=None
                                    , index_col = None
                                    , usecols= 'A:R'
                                    , engine= 'openpyxl'
                                    , dtype={col: str for col in Base_Texto})
        df_bdKilosVal = pd.DataFrame(bdKilosVal)
        df_ordenado = df_bdKilosVal.copy()
        df_ordenado = df_ordenado.rename(columns={'KILO_NETO': 'KILO_PROYECTADO','AÑO': 'AÑO_VAL','MES': 'MES_VAL'})
        df_ordenado.to_excel('C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenProyectado.xlsx'
            , index=False
            , sheet_name='Hoja1'                                               
            , engine='openpyxl')
#----
        print('\n---------------------------------\n')
        break
    elif BDCATEGORIA in FILES:
        print('\n---------------------------------\n')
        print(f" El archivo {BDCATEGORIA} está presente. Se realizará la explosión en función para Categoría.")
        bdCategoria = pd.read_excel(
            io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenDeProyección/bdCategoria.xlsx'
                                    , sheet_name='Hoja1'
                                    , header=0
                                    , names=None
                                    , index_col = None
                                    , usecols= 'A:D'
                                    , engine= 'openpyxl'
                                    , dtype={col: str for col in Base_Texto})
        df_bdCategoria = pd.DataFrame(bdCategoria)
        df_bdCategoria.columns = ['NOMBRE_CATEGORIA'
                                , 'CATEGORIA_VAL'
                                , 'MES_VAL'
                                , 'AÑO_VAL']
        #---- Fusionar df_agrupado1 con la data de Categoria
        df_bdCategoria = pd.merge(df_agrupado1
                                , df_bdCategoria
                                , on='NOMBRE_CATEGORIA')
        #---- Calcular la nueva columna con el porcentaje del total 'KILO_NETO'
        df_bdCategoria['KILO_PROYECTADO'] = (df_bdCategoria['PORC_TOTALCATEGORIA']
                                                    * df_bdCategoria['CATEGORIA_VAL'])
        df_bdCategoria = df_bdCategoria.rename(columns={'KILO_NETO': 'KILO_HISTORICO'})
        #--- Ordenamiento de Columnas
        Column_OrderVol1 = ['AÑO_VAL','MES_VAL','CODI_REGIONAL'
                            , 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                            , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                            , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                            , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                            , 'NOMBRE_TIPOOFERTA']
        Column_OrderVol2 = ['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                            , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                            , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                            , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                            , 'NOMBRE_TIPOOFERTA','AÑO_VAL', 'MES_VAL'
                            , 'KILO_HISTORICO','KILO_PROYECTADO']
        df_bdCategoria = df_bdCategoria[Column_OrderVol2]
        df_ordenado = df_bdCategoria.sort_values(by=Column_OrderVol1, ascending=True)
        # ---------------------------------------
        # La parte [col for col in df_ordenado.columns
        # if col not in Column_OrderVol2] asegura que cualquier columna
        # no incluida en Column_OrderVol2 se mantenga al final del DataFrame.
        #----------------------------------------------
        df_ordenado = df_ordenado[Column_OrderVol2 + [col for col in df_ordenado.columns if col not in Column_OrderVol2]]
        #----
        df_ordenado.to_excel('C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenProyectado.xlsx'
                    , index=False
                    , sheet_name='Hoja1'
                    , engine='openpyxl')
        #----
        print("** El archivo VolumenProyectado.xlsx se genero de manera exitosa y lo podrás visualizar en la ruta. **")
        print('\n---------------------------------\n')
        #print(df_agrupado1)
        #print(df_bdCategoria)
        #print(dfExplosionCat)
        break
    elif BDCANAL in FILES:
        print('\n---------------------------------\n')
        print(f" El archivo {BDCANAL} está presente. Se realizará la explosión en función para Categoría.")
        bdCanal = pd.read_excel(
            io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenDeProyección/bdCanal.xlsx'
                                    , sheet_name='Hoja1'
                                    , header=0
                                    , names=None
                                    , index_col = None
                                    , usecols= 'A:D'
                                    , engine= 'openpyxl'
                                    , dtype={col: str for col in Base_Texto})
        df_bdCanal = pd.DataFrame(bdCanal)
        df_bdCanal.columns = ['CODI_CANALAGRUP'
                                , 'CANAL_VAL'
                                , 'MES_VAL'
                                , 'AÑO_VAL']
        #---- Fusionar df_agrupado1 con la data de Categoria
        df_bdCanal = pd.merge(df_agrupado1
                                , df_bdCanal
                                , on='CODI_CANALAGRUP')
        #---- Calcular la nueva columna con el porcentaje del total 'KILO_NETO'
        df_bdCanal['KILO_PROYECTADO'] = (df_bdCanal['PORC_TOTALCATEGORIA']
                                                    * df_bdCanal['CANAL_VAL'])
        df_bdCanal = df_bdCanal.rename(columns={'KILO_NETO': 'KILO_HISTORICO'})
        #--- Ordenamiento de Columnas
        Column_OrderVol1 = ['AÑO_VAL','MES_VAL','CODI_REGIONAL'
                    , 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                    , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                    , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                    , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                    , 'NOMBRE_TIPOOFERTA']
        Column_OrderVol2 = ['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                            , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                            , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                            , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                            , 'NOMBRE_TIPOOFERTA','AÑO_VAL', 'MES_VAL'
                            , 'KILO_HISTORICO','KILO_PROYECTADO']
        df_bdCanal = df_bdCanal[Column_OrderVol2]
        df_ordenado = df_bdCanal.sort_values(by=Column_OrderVol1, ascending=True)
        # ---------------------------------------
        # La parte [col for col in df_ordenado.columns
        # if col not in Column_OrderVol2] asegura que cualquier columna
        # no incluida en Column_OrderVol2 se mantenga al final del DataFrame.
        #----------------------------------------------
        df_ordenado = df_ordenado[Column_OrderVol2 + [col for col in df_ordenado.columns if col not in Column_OrderVol2]]
        #----
        df_ordenado.to_excel('C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenProyectado.xlsx'
                    , index=False
                    , sheet_name='Hoja1'
                    , engine='openpyxl')
        #----
        print("** El archivo VolumenProyectado.xlsx se genero de manera exitosa y lo podrás visualizar en la ruta. **")
        print('\n---------------------------------\n')
        #print(df_agrupado1)
        #print(df_bdCategoria)
        #print(dfExplosionCat)
        break
    else:
        print('\n---------------------------------\n')
        print(" No hay archivos para realizar la explosión.")
        # Solicitar al usuario que presione Enter para volver a verificar
        input(" Presiona Enter cuando hayas agregado el archivo y quieras volver a verificar.")
# ---------------------------------------------


# ----------------------------------------------
# Segmento 8: Control de historico con N cantidad de meses
# para calcula la proyección de ventas
# ----------------------------------------------
#----  Copiar data del dataframe de VentasReal
df_HistVentas = df_BaseVentasReal.copy()
# Convertir las columnas 'año' y 'mes' a tipo str y luego a tipo int
df_HistVentas['Año'] = df_HistVentas['Año'].astype(str).astype(int)
df_HistVentas['Mes'] = df_HistVentas['Mes'].astype(str).astype(int)
# Calcular la fecha ficticia como el primer día de cada mes
df_HistVentas['fecha'] = pd.to_datetime(df_HistVentas['Año'].astype(str) + '-' + df_HistVentas['Mes'].astype(str) + '-01')
#----
# Variable para almacenar el número de mes
MES_HISTVENTAS = None
# Bucle para capturar un mes valido
while MES_HISTVENTAS is None:
    try:
        # Solicitar al usuario que ingrese el mes
        numero = int(input(" Ingrese la cantidad de meses del historico para Valorizar - "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= numero <= 12:
            MES_HISTVENTAS = numero  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese un número en el rango del 1 al 12.")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")
# Imprimir el número válido
print("\n Se Calculará para los Ultimos: ", MES_HISTVENTAS, " Meses")
#----
# Calcular la fecha de los ultimos N meses desde la fecha máxima
fecha_maximaVentas = df_HistVentas['fecha'].max()
fecha_limiteVentas = fecha_maximaVentas - relativedelta(months=(MES_HISTVENTAS)-1)
# Filtrar el DataFrame por los últimos 3 meses
df_filtradoVentas = df_HistVentas[df_HistVentas['fecha'] >= fecha_limiteVentas].copy()
# Eliminar la columna 'fecha' si no es necesaria
df_filtradoVentas = df_filtradoVentas.sort_values('fecha', ascending=False)
df_filtradoVentas = df_filtradoVentas.drop(columns=['fecha'])
# Imprimir el DataFrame resultante para validar que se tomo el historico correcto
df_filtradoVentas.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/Validación/ValidaciónMesVentas.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
# ----------------------------------------------


# ----------------------------------------------
# Segmento 9: Agrupación del Historico de Ventas
# ----------------------------------------------
# Columnas por las cuales se filtrará el Historico de Ventas Reales
col_tbd_Ventas = ['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                            , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                            , 'SUBGRUPO', 'COD ARTICULO','ARTICULO', 'MARCA'
                            , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA'
                            , 'NOMBRE_TIPOOFERTA','KILO_NETO'
                            , 'VALORTOTAL_PRODREGULAR', 'VALOR_DESC'
                            , 'VALOR_DEVO', 'CARTERA' ,'DESC_POSV'
                            , 'VALOR_INVERSI_PROMOC', 'EXTRAC']
df_tbd_Ventas = df_filtradoVentas[col_tbd_Ventas]
# Columnas por las cuales agrupar
col_AgrupVentas = ['CODI_REGIONAL', 'CODI_CANALAGRUP', 'CODI_CANALSUPE'
                    , 'CLIENTE', 'NOMBRE_CATEGORIA', 'DESC_GRUPO'
                    , 'SUBGRUPO', 'COD ARTICULO', 'MARCA'
                    , 'PRESENTACIÓN', 'TAMAÑO', 'TIPO_OFERTA']
#----
# Agrupar el DataFrame
# Puedes cambiar 'sum' a otras funciones de agregación
# (por ejemplo, 'mean' para promedio, 'count' para contar, etc.)
df_agrupadoVentas = df_tbd_Ventas.groupby(col_AgrupVentas).agg({
    'KILO_NETO': 'sum',
    'VALORTOTAL_PRODREGULAR': 'sum',
    'VALOR_DESC': 'sum',
    'VALOR_DEVO': 'sum',
    'CARTERA': 'sum',
    'DESC_POSV': 'sum',
    'VALOR_INVERSI_PROMOC': 'sum',
    'EXTRAC': 'sum'
}).reset_index()
# ----------------------------------------------


df_agrupadoVentas['Validación'] = np.where(
    (df_agrupadoVentas['VALORTOTAL_PRODREGULAR'] == 0) & (df_agrupadoVentas['KILO_NETO'] != 0),
    'VERIFICAR',
    'OK')

# ----------------------------------------------
# Segmento 10: Creación tabla para Valorizar
# ----------------------------------------------
# Columnas calculas agragadas al dataframe
#df_agrupadoVentas['PRODREGULAR X KILO'] = (df_agrupadoVentas['VALORTOTAL_PRODREGULAR']
# --- 'PRODREGULAR X KILO'
df_agrupadoVentas['PRODREGULAR X KILO'] = np.where(df_agrupadoVentas['VALORTOTAL_PRODREGULAR'] != 0
                                                   , df_agrupadoVentas['VALORTOTAL_PRODREGULAR'] 
                                                   / df_agrupadoVentas['KILO_NETO'], 0)
df_agrupadoVentas['PRODREGULAR X KILO'].replace(-np.inf, 0, inplace=True)
# --- 'VALOR_DESC %'
df_agrupadoVentas['VALOR_DESC %'] = np.where(df_agrupadoVentas['VALOR_DESC'] != 0
                                                   , df_agrupadoVentas['VALOR_DESC']
                                                   / df_agrupadoVentas['VALORTOTAL_PRODREGULAR'], 0)
df_agrupadoVentas['VALOR_DESC %'].replace(-np.inf, 0, inplace=True)
# --- 'VALOR_DEVO %'
df_agrupadoVentas['VALOR_DEVO %'] = np.where(df_agrupadoVentas['VALOR_DEVO'] != 0
                                                   , df_agrupadoVentas['VALOR_DEVO']
                                                   / df_agrupadoVentas['VALORTOTAL_PRODREGULAR'], 0)
df_agrupadoVentas['VALOR_DEVO %'].replace(-np.inf, 0, inplace=True)
# --- 'CARTERA %'
df_agrupadoVentas['CARTERA %'] = np.where(df_agrupadoVentas['CARTERA'] != 0
                                                   , df_agrupadoVentas['CARTERA']
                                                   / df_agrupadoVentas['VALORTOTAL_PRODREGULAR'], 0)
df_agrupadoVentas['CARTERA %'].replace(-np.inf, 0, inplace=True)
# --- 'DESC_POSV %'
df_agrupadoVentas['DESC_POSV %'] = np.where(df_agrupadoVentas['DESC_POSV'] != 0
                                                   , df_agrupadoVentas['DESC_POSV']
                                                   / df_agrupadoVentas['VALORTOTAL_PRODREGULAR'], 0)
df_agrupadoVentas['DESC_POSV %'].replace(-np.inf, 0, inplace=True)
# --- 'VALOR_INVERSI_PROMOC %''
df_agrupadoVentas['VALOR_INVERSI_PROMOC %'] = np.where(df_agrupadoVentas['VALOR_INVERSI_PROMOC'] != 0
                                                   , df_agrupadoVentas['VALOR_INVERSI_PROMOC']
                                                   / df_agrupadoVentas['VALORTOTAL_PRODREGULAR'], 0)
df_agrupadoVentas['VALOR_INVERSI_PROMOC %'].replace(-np.inf, 0, inplace=True)
# --- 'EXTRAC %'
df_agrupadoVentas['EXTRAC %'] = np.where(df_agrupadoVentas['EXTRAC'] != 0
                                                   , df_agrupadoVentas['EXTRAC']
                                                   / df_agrupadoVentas['VALORTOTAL_PRODREGULAR'], 0)
df_agrupadoVentas['EXTRAC %'].replace(-np.inf, 0, inplace=True)
# ----------------------------------------------


# ----------------------------------------------
# Segmento 11: Cruce dos Dataframe Principales (Volumen/Ventas)
# ----------------------------------------------
# Crear una columna llave
df_agrupadoVentas['LLAVE'] = df_agrupadoVentas.apply(lambda row: '_'.join(map(str, row[col_AgrupVentas])), axis=1)
# Imprimir el DataFrame resultante para validar que se tomo el historico correcto
df_agrupadoVentas.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/Validación/AgrupaciónFinalVentas.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
# ----------------------------------------------
# Copiar data del dataframe del VolumenProyectado
df_agrupadoVolumen = df_ordenado.copy()
# Crear una columna llave
df_agrupadoVolumen['LLAVE'] = df_agrupadoVolumen.apply(lambda row: '_'.join(map(str, row[col_AgrupVentas])), axis=1)
df_agrupadoVolumen.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/Validación/AgrupaciónFinalVolumen.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
print('\n---------------------------------\n')
# Fusionar VolumenProyectado con la AgrupaciónDeVentas
df_Valorizacion = pd.merge(df_agrupadoVolumen
                        , df_agrupadoVentas
                        , on='LLAVE', how='inner')
# ----------------------------------------------
df_Valorizacion = df_Valorizacion.rename(columns={'NOMBRE_CATEGORIA_x': 'NOMBRE_CATEGORIA'
                                                  , 'CODI_CANALSUPE_x': 'CODI_CANALSUPE'
                                                  , 'CLIENTE_x': 'CLIENTE'
                                                  , 'DESC_GRUPO_x': 'DESC_GRUPO'
                                                  , 'SUBGRUPO_x': 'SUBGRUPO'
                                                  , 'CODI_CANALAGRUP_x': 'CODI_CANALAGRUP'
                                                  , 'CODI_REGIONAL_x': 'CODI_REGIONAL'
                                                  , 'COD ARTICULO_x': 'COD ARTICULO'
                                                  , 'MARCA_x': 'MARCA'
                                                  , 'PRESENTACIÓN_x': 'PRESENTACIÓN'
                                                  , 'TAMAÑO_x': 'TAMAÑO'
                                                  , 'TIPO_OFERTA_x': 'TIPO_OFERTA'})
# ----------------------------------------------
#df_Valorizacion = pd.concat([df_agrupadoVolumen,Columns_Ventas], axis=1)
df_Valorizacion.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/Validación/AgrupaciónFinalVentas+Volumen.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
# ----------------------------------------------


# ----------------------------------------------
# Segmento 12: Lectura y cruce de Archivos Editores de la Proyección
# ----------------------------------------------
tbIncrementos = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/ModificadoresProyección/ModificadoresProyección.xlsx'
                            , sheet_name='Precio'
                            , header=0
                            , names=None
                            , index_col = None
                            , usecols= 'A:F'
                            , engine= 'openpyxl'
                            #, dtype={col: str for col in Base_Texto}
                            )
db_Incrementos = pd.DataFrame(tbIncrementos)
# ----
tbDescuentos = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/ModificadoresProyección/ModificadoresProyección.xlsx'
                            , sheet_name='Descuento'
                            , header=0
                            , names=None
                            , index_col = None
                            , usecols= 'A:E'
                            , engine= 'openpyxl'
                            #, dtype={col: str for col in Base_Texto}
                            )
df_Descuentos = pd.DataFrame(tbDescuentos)
# ----
tbOfertacion = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/ModificadoresProyección/ModificadoresProyección.xlsx'
                            , sheet_name='Ofertación'
                            , header=0
                            , names=None
                            , index_col = None
                            , usecols= 'A:E'
                            , engine= 'openpyxl'
                            #, dtype={col: str for col in Base_Texto}
                            )
db_Ofertacion = pd.DataFrame(tbOfertacion)
# ----
tbExtra = pd.read_excel(
    io= 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/ModificadoresProyección/ModificadoresProyección.xlsx'
                            , sheet_name='ExtraContenido'
                            , header=0
                            , names=None
                            , index_col = None
                            , usecols= 'A:E'
                            , engine= 'openpyxl'
                            #, dtype={col: str for col in Base_Texto}
                            )
db_tbExtra = pd.DataFrame(tbExtra)
# ----------------------------------------------
# Cruzar Incrementos con VentasProyectadas
df_Valorizacion = pd.merge(df_Valorizacion
                        , db_Incrementos
                        , on=['NOMBRE_CATEGORIA', 'DESC_GRUPO', 'CODI_CANALSUPE', 'CLIENTE', 'CODI_REGIONAL']
                        , how='inner')
df_Valorizacion = df_Valorizacion.rename(columns={'PORCENTAJE': 'INCREMENTO'})
# Cruzar Descuentos con VentasProyectadas
df_Valorizacion = pd.merge(df_Valorizacion
                        , df_Descuentos
                        , on=['NOMBRE_CATEGORIA', 'DESC_GRUPO', 'CODI_CANALSUPE', 'CLIENTE']
                        , how='inner')
df_Valorizacion = df_Valorizacion.rename(columns={'PORCENTAJE': 'DESCUENTO'})
# ----
# Cruzar Ofertación con VentasProyectadas
df_Valorizacion = pd.merge(df_Valorizacion
                        , db_Ofertacion
                        , on=['NOMBRE_CATEGORIA', 'DESC_GRUPO', 'CODI_CANALSUPE', 'CLIENTE']
                        , how='inner')
df_Valorizacion = df_Valorizacion.rename(columns={'PORCENTAJE': 'OFERTACION'})
# ----
# Cruzar ExtraContenido con VentasProyectadas
df_Valorizacion = pd.merge(df_Valorizacion
                        , db_tbExtra
                        , on=['NOMBRE_CATEGORIA', 'DESC_GRUPO', 'CODI_CANALSUPE', 'CLIENTE']
                        , how='inner')
df_Valorizacion = df_Valorizacion.rename(columns={'PORCENTAJE': 'EXTRA'})
# ----------------------------------------------


# ----------------------------------------------
# Segmento 13: Valorización Ventas Proyectadas
# ----------------------------------------------
# ---- Gestión Columnas ValorTotal_ProdRegular Proyectado
# ----------------------------------------------
# Variable para respuesta
PRECIOREGULAR = None
# Bucle para capturar una respuesta valida
while PRECIOREGULAR is None:
    try:
        # Solicitar al usuario que escoja una opción
        rpta_Inc = int(input(" ¿Quieres modificar el Precio? \n Si = 1 \n No = 2  \n Ingresa un Número: "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= rpta_Inc <= 2:
            PRECIOREGULAR = rpta_Inc  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese 1 o 2 según quieras realizar la proyección")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")
#----- Función para aplicar la condición
def fn_IncrementoPrecio(valor):
    if valor == 1:
        return ((df_Valorizacion['KILO_PROYECTADO'] * df_Valorizacion['PRODREGULAR X KILO'])*df_Valorizacion['INCREMENTO'])+(df_Valorizacion['KILO_PROYECTADO'] * df_Valorizacion['PRODREGULAR X KILO'])
    else:
        return df_Valorizacion['KILO_PROYECTADO'] * df_Valorizacion['PRODREGULAR X KILO']
print('\n -----------------------------\n')
#------ Aplicar la función y crear una nueva columna basada en la condición
df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] = fn_IncrementoPrecio(PRECIOREGULAR)
# ----------------------------------------------
df_Valorizacion['VALOR_DESC_PROY'] = (df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY']
                                            * df_Valorizacion['VALOR_DESC %'])
df_Valorizacion['VALOR_DEVO_PROY'] = (df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY']
                                            * df_Valorizacion['VALOR_DEVO %'])
df_Valorizacion['CARTERA_PROY'] = (df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY']
                                            * df_Valorizacion['CARTERA %'])
# ----------------------------------------------
# ---- Gestión Columnas Descuento Proyectado
# ----------------------------------------------
# Variable para respuesta
DESCUENTO = None
# Bucle para capturar una respuesta valida
while DESCUENTO is None:
    try:
        # Solicitar al usuario que escoja una opción
        rpta_Desc = int(input(" ¿Quieres modificar el Descuento? \n Si = 1 \n No = 2  \n Ingresa un Número: "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= rpta_Desc <= 2:
            DESCUENTO = rpta_Desc  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese 1 o 2 según quieras realizar la proyección")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")
#----- Función para aplicar la condición
def fn_IncrementoDesc(valor):
    if valor == 1:
        return ((df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] * df_Valorizacion['DESC_POSV %'])*df_Valorizacion['DESCUENTO'])+(df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] * df_Valorizacion['DESC_POSV %'])
    else:
        return df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] * df_Valorizacion['DESC_POSV %']
print('\n | -----------------------------\n')
#------ Aplicar la función y crear una nueva columna basada en la condición
df_Valorizacion['DESC_POSV_PROY'] = fn_IncrementoDesc(DESCUENTO)
# ----------------------------------------------
# ----------------------------------------------
# ---- Gestión Columnas Ofertación Proyectado
# ----------------------------------------------
# Variable para respuesta
OFERTACION = None
# Bucle para capturar una respuesta valida
while OFERTACION is None:
    try:
        # Solicitar al usuario que escoja una opción
        rpta_Ofer = int(input(" ¿Quieres modificar la Ofertación? \n Si = 1 \n No = 2  \n Ingresa un Número: "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= rpta_Ofer <= 2:
            OFERTACION = rpta_Ofer  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese 1 o 2 según quieras realizar la proyección")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")
#----- Función para aplicar la condición
def fn_IncrementoOfert(valor):
    if valor == 1:
        return ((df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] * df_Valorizacion['VALOR_INVERSI_PROMOC %'])*df_Valorizacion['OFERTACION'])+(df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] * df_Valorizacion['VALOR_INVERSI_PROMOC %'])
    else:
        return df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] * df_Valorizacion['VALOR_INVERSI_PROMOC %']
print('\n -----------------------------\n')
#------ Aplicar la función y crear una nueva columna basada en la condición
df_Valorizacion['VALOR_INVERSI_PROMOC_PROY'] = fn_IncrementoOfert(OFERTACION)
# ----------------------------------------------
# Gestión Columnas ExtraContenido 0 u Original
# ----------------------------------------------
# Variable para respuesta
EXTRA_CONT = None
# Bucle para capturar una respuesta valida
while EXTRA_CONT is None:
    try:
        # Solicitar al usuario que escoja una opción
        rpta_Ext = int(input(" ¿Proyección con Extra Contenido o en 0?  \n Si = 1 \n No = 2  \n Ingresa un Número: "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= rpta_Ext <= 2:
            EXTRA_CONT = rpta_Ext  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese 1 o 2 según quieras realizar la proyección")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")

# Función para aplicar la condición
def fn_ExtraContenido(valor):
    if valor == 1:
        return df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY'] * df_Valorizacion['EXTRAC %']
    else:
        return 0
# Aplicar la función y crear una nueva columna basada en la condición
df_Valorizacion['EXTRAC_PROY1'] = fn_ExtraContenido(EXTRA_CONT)
# ----------------------------------------------
# ----  Gestión Columnas ExtraContenido
# ----------------------------------------------
# Variable para respuesta
EXTRACTN = None
# Bucle para capturar una respuesta valida
while EXTRACTN is None:
    try:
        # Solicitar al usuario que escoja una opción
        rpta_ExtCtn = int(input(" ¿Quieres modificar el Extra Contenido? \n Si = 1 \n No = 2  \n Ingresa un Número: "))
        # Validar que el número esté en el rango de 1 a 12
        if 1 <= rpta_ExtCtn <= 2:
            EXTRACTN = rpta_ExtCtn  # Almacenar el número si es válido
        else:
            print(" Por favor, ingrese 1 o 2 según quieras realizar la proyección")
    except ValueError:
        print(" Por favor, ingrese un valor numérico válido.")
#----- Función para aplicar la condición
def fn_IncrementoExtra(valor):
    if valor == 1:
        return (df_Valorizacion['EXTRAC_PROY1']*df_Valorizacion['EXTRA'])+(df_Valorizacion['EXTRAC_PROY1'])
    else:
        return df_Valorizacion['EXTRAC_PROY1']
print('\n -----------------------------\n')
#------ Aplicar la función y crear una nueva columna basada en la condición
df_Valorizacion['EXTRAC_PROY'] = fn_IncrementoExtra(EXTRACTN)
# ----------------------------------------------
df_Valorizacion['VALOR_NETO_PROY'] = (df_Valorizacion['VALORTOTAL_PRODREGULAR_PROY']
                                      + df_Valorizacion['VALOR_DESC_PROY']
                                      + df_Valorizacion['VALOR_DEVO_PROY']
                                      + df_Valorizacion['CARTERA_PROY']
                                      + df_Valorizacion['DESC_POSV_PROY']
                                      + df_Valorizacion['VALOR_INVERSI_PROMOC_PROY']
                                      + df_Valorizacion['EXTRAC_PROY'])
# ----------------------------------------------
#Columns_Valorizacion = df_Valorizacion[['LLAVE','VALORTOTAL_PRODREGULAR_PROY', 'VALOR_DESC_PROY'
#                                        ,'VALOR_DEVO_PROY','CARTERA_PROY','DESC_POSV_PROY'
#                                        ,'VALOR_INVERSI_PROMOC_PROY','EXTRAC_PROY']]
# ----------------------------------------------


# ----------------------------------------------
# Segmento 14: Eliminación de Columnas
# ----------------------------------------------
#'KILO_HISTORICO',
Drop_Columns = df_Valorizacion[[ 'LLAVE','NOMBRE_CATEGORIA_y'
                                , 'CODI_CANALSUPE_y', 'CLIENTE_y'
                                , 'VALORTOTAL_PRODREGULAR'
                                , 'DESC_GRUPO_y', 'SUBGRUPO_y'
                                , 'CODI_CANALAGRUP_y', 'CODI_REGIONAL_y', 'COD ARTICULO_y'
                                , 'MARCA_y', 'PRESENTACIÓN_y', 'TAMAÑO_y', 'TIPO_OFERTA_y'
                                , 'VALOR_DESC', 'VALOR_DEVO'
                                , 'CARTERA', 'DESC_POSV', 'VALOR_INVERSI_PROMOC'
                                , 'EXTRAC', 'PRODREGULAR X KILO', 'VALOR_DESC %'
                                , 'VALOR_DEVO %', 'CARTERA %', 'DESC_POSV %'
                                , 'VALOR_INVERSI_PROMOC %', 'EXTRAC %', 'KILO_NETO'
                                , 'DESCUENTO', 'OFERTACION', 'INCREMENTO','EXTRA','EXTRAC_PROY1']]
# Seleccionar las columnas que no están en la lista de exclusiones
columnas_a_mantener = [col for col in df_Valorizacion.columns if col not in Drop_Columns]
df_Valorizacion = df_Valorizacion[columnas_a_mantener]
# ----------------------------------------------


# ----------------------------------------------
# Segmento 15: Exportación del archivo final de Proyecciones
# ----------------------------------------------
df_Valorizacion.to_excel(
    'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VentasProyectadas.xlsx'
    , index=False
    , sheet_name='Hoja1'
    , engine='openpyxl')
print("\n ** Se ha generado el archivo final de proyecciones con éxitos ** ")
# ----------------------------------------------


# ----------------------------------------------
# Segmento 16: Diseño Visual del Excel con openpyxl
# ----------------------------------------------
# Ruta al archivo
archivo_xlsx = 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VolumenProyectado.xlsx'
# Cargar el libro de trabajo
wb = load_workbook(archivo_xlsx)
hoja = wb['Hoja1']

# Aplica formato al encabezado (primera fila)
for cell in hoja[1]:
    cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)

# Ajustar el ancho de las columnas al contenido
for columna in hoja.columns:
    max_length = 0
    columna = [cell for cell in columna]
    for cell in columna:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 4)
    hoja.column_dimensions[columna[0].column_letter].width = adjusted_width
# Guarda el archivo Excel
wb.save(archivo_xlsx)
# Cierra el libro de trabajo
wb.close()
# ----------------------------------------------
# Ruta al archivo
archivo_xlsx = 'C:/Users/cristian.segura/Documents/Python/Proyecto2/Proyección de Ventas/VentasProyectadas.xlsx'
# Cargar el libro de trabajo
wb = load_workbook(archivo_xlsx)
hoja = wb['Hoja1']

# Aplica formato al encabezado (primera fila)
for cell in hoja[1]:
    cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)

# Ajustar el ancho de las columnas al contenido
for columna in hoja.columns:
    max_length = 0
    columna = [cell for cell in columna]
    for cell in columna:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 4)
    hoja.column_dimensions[columna[0].column_letter].width = adjusted_width
# Guarda el archivo Excel
wb.save(archivo_xlsx)
# Cierra el libro de trabajo
wb.close()
# ----------------------------------------------


# ----------------------------------------------
# Segmento 17: Tiempos de ejecución
# ----------------------------------------------
# Calcular la duración total de la ejecución
tiempo_fin = time.time()
# Calcular la duración total de la ejecución en segundos
duracion_total_segundos = tiempo_fin - tiempo_inicio
# Obtener horas, minutos y segundos
horas, rem = divmod(duracion_total_segundos, 3600)
minutos, segundos = divmod(rem, 60)
# Imprimir el tiempo total de ejecución en formato hh/mm/ss
print(f' Tiempo total de ejecución: {int(horas):02d}:{int(minutos):02d}:{int(segundos):02d}')
#time.sleep(10)
input(" Presiona Enter para salir")
# ----------------------------------------------